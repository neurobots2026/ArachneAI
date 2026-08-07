from io import BytesIO

from openpyxl import Workbook


def generate_fake_payroll_xlsx(org_name: str, department: str, marker: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    ws.append(["Employee", "Department", "Salary", "Tracking"])
    ws.append(["Jane Doe", department, 85000, marker])
    ws.append(["John Smith", department, 92000, f"ref_{marker}"])
    ws["A1"].comment = None
    ws.cell(row=1, column=1, value=f"{org_name} - Confidential")
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
